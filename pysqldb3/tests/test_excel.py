from os import path
from ..excel import DbConnectExcel
from . import TestHelpers
import pytest
import pandas as pd
import numpy as np
import openpyxl as opxl

testing_schema = 'testing'
default_multi_sheetnames = ['Sheet1','Sheet2','Sheet3']

class TestExcel:

    def test_df_to_excel(self):
        # Create dataframe with random ints and save it
        excel_path = TestHelpers.test_data_folder_path('testexcelsingle.xlsx')
        test_df = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        DbConnectExcel.df_to_excel(self, df=test_df, filename='testexcelsingle')

        # Assert existance and xlsx format
        assert path.exists(excel_path)
        assert path.splitext(excel_path)[1] == 'xlsx'

        # Open excel_path as a new df
        test_df_2 = pd.read_excel(excel_path)

        # Assert dfs are equal
        assert test_df_2.equals(test_df)
    
    def test_df_to_excel_xls(self):
        # Create dataframe with random ints and save it
        excel_path = TestHelpers.test_data_folder_path('testexcelsingle.xls')
        test_df = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        DbConnectExcel.df_to_excel(self, df=test_df, filename='testexcelsingle', as_xls=True)

        # Assert existance and xls format
        assert path.exists(excel_path)
        assert path.splitext(excel_path)[1] == 'xls'

        # Open excel_path as a new df
        test_df_2 = pd.read_excel(excel_path)

        # Assert dfs are equal
        assert test_df_2.equals(test_df)
    
    def test_df_to_excel_custom_tab_name(self):
        # Create dataframe with random ints and save it
        excel_path = TestHelpers.test_data_folder_path('testexcelsinglenamedtab.xlsx')
        test_df = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        DbConnectExcel.df_to_excel(self, df=test_df, filename='testexcelsingle', tab_name='customtabname')

        # Assert existance
        assert path.exists(excel_path)

        # Open excel_path as a new df
        test_df_2 = pd.read_excel(excel_path)

        # Assert dfs are equal
        assert test_df_2.equals(test_df)
    
    def test_df_to_excel_multiple_tabs(self):
        excel_path = TestHelpers.test_data_folder_path('testexcelmulti.xlsx')

        # Create random dfs
        test_df1 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        test_df2 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        test_df3 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))

        # Save
        DbConnectExcel.df_to_excel_multi(self, df_list=[test_df1,test_df2,test_df3], filename='testexcelmulti')

        # Assert existence
        assert path.exists(excel_path)

        # Assert proper tabs
        wb = opxl.load_workbook(excel_path, read_only=True)
        assert wb.sheetnames == []
        wb.close()

    def test_df_to_excel_multiple_tabs_diff_df_and_name_lengths(self):
        excel_path = TestHelpers.test_data_folder_path('testexcelmultinamedtabs.xlsx')

        # Create random dfs
        test_df1 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        test_df2 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))
        test_df3 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(5,5)))

        # Try to save but assert ValueError
        with pytest.raises(ValueError):
            DbConnectExcel.df_to_excel_multi(self, df_list=[test_df1,test_df2,test_df3], filename='testexceldiffdfnamelength',
            tab_name_list=['tab1','tab2'])

    def test_pg_to_excel(self):
        pg = TestHelpers.get_pg_dbc_instance()
        excel_path = TestHelpers.test_data_folder_path('testpgtoexcel.xlsx')

        # create random df
        test_df = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))

        # Insert into DB
        pg.connect()
        table_name = f'test_pg_to_excel_{pg.user}'
        pg.dataframe_to_table(test_df, table=table_name)
        assert pg.table_exists(table_name)

        # Export and assert existence
        DbConnectExcel.df_to_excel(self, df=test_df, filename='testpgtoexcel')
        assert path.exists(TestHelpers.test_data_folder_path('testpgtoexcel.xlsx'))

        # cleanup
        pg.drop_table(schema=testing_schema, table=table_name)


    def test_pg_to_excel_multiple_tabs(self):
        pg = TestHelpers.get_pg_dbc_instance()
        excel_path = TestHelpers.test_data_folder_path('testpgtoexcelmulti.xlsx')
        
        # Create random dfs
        test_df1 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df2 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df3 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))

        # Insert into DB
        pg.connect()
        pg.dataframe_to_table(test_df1, table=f'pdb3_test_pg_xlsx_df1_{pg.user}')
        pg.dataframe_to_table(test_df2, table=f'pdb3_test_pg_xlsx_df2_{pg.user}')
        pg.dataframe_to_table(test_df3, table=f'pdb3_test_pg_xlsx_df3_{pg.user}')
        assert pg.table_exists(f'pdb3_test_pg_xlsx_df1_{pg.user}') and pg.table_exists(f'pdb3_test_pg_xlsx_df2_{pg.user}') \
            and pg.table_exists(f'pdb3_test_pg_xlsx_df3_{pg.user}')

        # Export and assert existence
        DbConnectExcel.df_to_excel_multi(self, df_list=[test_df1,test_df2,test_df3],filename='testpgtoexcelmulti')
        assert path.exists(excel_path)

        # Assert correct sheetnames
        xlsx = opxl.load_workbook(excel_path)
        assert xlsx.sheetnames == default_multi_sheetnames

        # Cleanup
        pg.drop_table(schema=testing_schema, table=f'pdb3_test_pg_xlsx_df1_{pg.user}')
        pg.drop_table(schema=testing_schema, table=f'pdb3_test_pg_xlsx_df2_{pg.user}')
        pg.drop_table(schema=testing_schema, table=f'pdb3_test_pg_xlsx_df3_{pg.user}')

    def test_pg_to_excel_multiple_tabs_custom_names(self):
        pg = TestHelpers.get_pg_dbc_instance()
        excel_path = TestHelpers.test_data_folder_path('testpgtoexcelmultinamed.xlsx')
        
        # Create random dfs
        test_df1 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df2 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df3 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        table_names = [f'pg_to_xlsx_1_{pg.user}',f'pg_to_xlsx_2_{pg.user}',f'pg_to_xlsx_3_{pg.user}']

        # Insert into DB
        pg.connect()
        pg.dataframe_to_table(test_df1, table=table_names[0])
        pg.dataframe_to_table(test_df2, table=table_names[1])
        pg.dataframe_to_table(test_df3, table=table_names[2])
        for name in table_names:
            assert pg.table_exists(name)

        # Export and assert existence
        DbConnectExcel.df_to_excel_multi(self, df_list=[test_df1,test_df2,test_df3],filename='testpgtoexcelmultinamed')
        assert path.exists(excel_path)

        # TODO: add checks for contents of excel files

        # Cleanup
        for i in len(table_names):
            pg.drop_table(schema='testing', table=table_names[i])

    def test_ms_to_excel(self):
        ms = TestHelpers.get_ms_dbc_instance()
        excel_path = TestHelpers.test_data_folder_path('testmstoexcel.xlsx')
        
        # create random df
        test_df = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))

        # Insert into DB
        ms.connect()
        table_name = f'test_ms_to_excel_{ms.user}'
        ms.dataframe_to_table(test_df, table=table_name)
        assert ms.table_exists(table_name)

        # Export and assert existence
        DbConnectExcel.df_to_excel(self, df=test_df, filename='testmstoexcel')
        assert path.exists(excel_path)

        # cleanup
        ms.drop_table(schema=testing_schema, table=table_name)

    def test_ms_to_excel_multiple_tabs(self):
        ms = TestHelpers.get_ms_dbc_instance()
        excel_path = TestHelpers.test_data_folder_path('testmstoexcelmulti.xlsx')
        
        # create random df's
        test_df1 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df2 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df3 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        
        # insert into DB
        ms.connect()
        table_names = [f'ms_to_xlsx_1_{ms.user}', f'ms_to_xlsx_2_{ms.user}', f'ms_to_xlsx_3_{ms.user}']
        ms.dataframe_to_table(test_df1, table=table_names[0])
        ms.dataframe_to_table(test_df2, table=table_names[1])
        ms.dataframe_to_table(test_df3, table=table_names[2])
        
        # export xlsx and assert existence
        DbConnectExcel.df_to_excel_multi(self, df_list=[test_df1,test_df2,test_df3], filename='testmstoexcelmulti')
        assert path.exists(excel_path)

        # cleanup
        ms.drop_table(table_names[0], schema=testing_schema)
        ms.drop_table(table_names[1], schema=testing_schema)
        ms.drop_table(table_names[2], schema=testing_schema)
    
    def test_ms_to_excel_multiple_tabs_custom_names(self):
        ms = TestHelpers.get_ms_dbc_instance()
        excel_path = TestHelpers.test_data_folder_path('testmstoexcelmultinamed.xlsx')
        
        # create random df's
        test_df1 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df2 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))
        test_df3 = pd.DataFrame(np.random.random_integers(low=0,high=999999,size=(10,10)))

        # insert into DB
        ms.connect()
        table_names = [f'ms_to_xlsx_1_{ms.user}',f'ms_to_xlsx_2_{ms.user}',f'ms_to_xlsx_3_{ms.user}']
        ms.dataframe_to_table(test_df1, table=table_names[0])
        ms.dataframe_to_table(test_df2, table=table_names[1])
        ms.dataframe_to_table(test_df3, table=table_names[2])

        # export xlsx and assert existence
        DbConnectExcel.df_to_excel_multi(self, df_list=[test_df1,test_df2,test_df3],filename='testmstoexcelmultinamed', tab_name_list=table_names)
        assert path.exists(excel_path)

        # cleanup
        ms.drop_table(table_names[0], schema=testing_schema)
        ms.drop_table(table_names[1], schema=testing_schema)
        ms.drop_table(table_names[2], schema=testing_schema)